"""
Data processing handlers.

Handlers for CSV, JSON, XML, Excel, PDF processing.
"""

import csv
import json
import xml.etree.ElementTree as ET
from xml.dom import minidom
import time
from io import StringIO, BytesIO
from typing import Dict, Any, List, Optional, Union
from pathlib import Path

from handlers.base import (
    TaskHandler,
    HandlerResult,
    HandlerConfidence,
)


class CSVParseHandler(TaskHandler):
    """Parse CSV files with automatic delimiter detection."""
    
    name = "csv_parse"
    category = "data"
    description = "Parse CSV files with automatic delimiter detection"
    requires_filesystem = False
    
    input_schema = {
        'content': {'type': 'string', 'description': 'CSV content (if not using file)'},
        'path': {'type': 'string', 'description': 'File path (alternative to content)'},
        'delimiter': {'type': 'string', 'description': 'Field delimiter (auto-detected if not specified)'},
        'has_header': {'type': 'boolean', 'default': True, 'description': 'First row is header'},
        'encoding': {'type': 'string', 'default': 'utf-8', 'description': 'File encoding'},
    }
    
    output_schema = {
        'headers': {'type': 'array', 'description': 'Column headers'},
        'rows': {'type': 'array', 'description': 'Data rows as dictionaries'},
        'row_count': {'type': 'integer', 'description': 'Number of rows'},
        'column_count': {'type': 'integer', 'description': 'Number of columns'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        task = params.get('task', '')
        if task in ['csv_parse', 'parse_csv', 'csv_read', 'csv']:
            return HandlerConfidence.PERFECT.value
        if 'csv' in task.lower() or params.get('path', '').endswith('.csv'):
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        
        try:
            # Get content
            if 'content' in params:
                content = params['content']
            elif 'path' in params:
                with open(params['path'], 'r', encoding=params.get('encoding', 'utf-8')) as f:
                    content = f.read()
            else:
                return HandlerResult(success=False, error="No content or path provided")
            
            # Detect delimiter
            delimiter = params.get('delimiter')
            if not delimiter:
                # Auto-detect: count common delimiters in first line
                first_line = content.split('\n')[0]
                counts = {
                    ',': first_line.count(','),
                    '\t': first_line.count('\t'),
                    ';': first_line.count(';'),
                    '|': first_line.count('|'),
                }
                delimiter = max(counts, key=counts.get)
            
            has_header = params.get('has_header', True)
            
            # Parse CSV
            reader = csv.reader(StringIO(content), delimiter=delimiter)
            rows = list(reader)
            
            if not rows:
                return HandlerResult(
                    success=True,
                    data={'headers': [], 'rows': [], 'row_count': 0, 'column_count': 0},
                )
            
            if has_header:
                headers = rows[0]
                data_rows = []
                for row in rows[1:]:
                    row_dict = {}
                    for i, val in enumerate(row):
                        key = headers[i] if i < len(headers) else f'col_{i}'
                        row_dict[key] = val
                    data_rows.append(row_dict)
            else:
                headers = [f'col_{i}' for i in range(len(rows[0]))]
                data_rows = [{headers[i]: val for i, val in enumerate(row)} for row in rows]
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'headers': headers,
                    'rows': data_rows,
                    'row_count': len(data_rows),
                    'column_count': len(headers),
                    'delimiter': delimiter,
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class CSVTransformHandler(TaskHandler):
    """Transform CSV data with filtering, mapping, aggregation."""
    
    name = "csv_transform"
    category = "data"
    description = "Transform CSV with filtering, mapping, aggregation"
    requires_filesystem = False
    
    input_schema = {
        'rows': {'type': 'array', 'required': True, 'description': 'Input rows'},
        'operations': {'type': 'array', 'required': True, 'description': 'List of operations'},
    }
    
    output_schema = {
        'rows': {'type': 'array', 'description': 'Transformed rows'},
        'row_count': {'type': 'integer', 'description': 'Output row count'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['csv_transform', 'transform_csv', 'csv_filter']:
            return HandlerConfidence.PERFECT.value
        if 'rows' in params and 'operations' in params:
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        rows = params['rows']
        operations = params['operations']
        
        try:
            result_rows = list(rows)  # Copy
            
            for op in operations:
                op_type = op.get('type') or op.get('op')
                
                if op_type == 'filter':
                    # Filter rows by condition
                    field = op['field']
                    operator = op.get('operator', '==')
                    value = op['value']
                    
                    def matches(row):
                        row_val = row.get(field)
                        if operator == '==':
                            return row_val == value
                        elif operator == '!=':
                            return row_val != value
                        elif operator == '>':
                            return row_val > value
                        elif operator == '<':
                            return row_val < value
                        elif operator == '>=':
                            return row_val >= value
                        elif operator == '<=':
                            return row_val <= value
                        elif operator == 'contains':
                            return value in str(row_val)
                        elif operator == 'in':
                            return row_val in value
                        return False
                    
                    result_rows = [r for r in result_rows if matches(r)]
                
                elif op_type == 'map':
                    # Add/transform columns
                    for row in result_rows:
                        for key, expr in op.get('fields', {}).items():
                            # Simple expression evaluation
                            if isinstance(expr, str) and expr.startswith('$'):
                                # Reference another field
                                ref_field = expr[1:]
                                row[key] = row.get(ref_field)
                            else:
                                row[key] = expr
                
                elif op_type == 'select':
                    # Select specific columns
                    columns = op.get('columns', [])
                    result_rows = [{k: v for k, v in r.items() if k in columns} for r in result_rows]
                
                elif op_type == 'rename':
                    # Rename columns
                    mapping = op.get('mapping', {})
                    for row in result_rows:
                        for old_key, new_key in mapping.items():
                            if old_key in row:
                                row[new_key] = row.pop(old_key)
                
                elif op_type == 'sort':
                    # Sort rows
                    key_field = op.get('key')
                    reverse = op.get('reverse', False)
                    result_rows.sort(key=lambda r: r.get(key_field, ''), reverse=reverse)
                
                elif op_type == 'limit':
                    # Limit rows
                    n = op.get('n', len(result_rows))
                    result_rows = result_rows[:n]
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'rows': result_rows,
                    'row_count': len(result_rows),
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class JSONTransformHandler(TaskHandler):
    """Transform JSON data with jq-like operations."""
    
    name = "json_transform"
    category = "data"
    description = "Transform JSON with jq-like path expressions"
    requires_filesystem = False
    
    input_schema = {
        'data': {'type': 'any', 'required': True, 'description': 'Input JSON data'},
        'path': {'type': 'string', 'description': 'Path expression (e.g., ".items[].name")'},
        'operations': {'type': 'array', 'description': 'Transform operations'},
    }
    
    output_schema = {
        'data': {'type': 'any', 'description': 'Transformed data'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['json_transform', 'transform_json', 'jq']:
            return HandlerConfidence.PERFECT.value
        if 'data' in params and ('path' in params or 'operations' in params):
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        data = params['data']
        path = params.get('path', '.')
        operations = params.get('operations', [])
        
        try:
            # Simple path traversal
            result = self._traverse_path(data, path)
            
            # Apply operations
            for op in operations:
                result = self._apply_operation(result, op)
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data=result,
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )
    
    def _traverse_path(self, data: Any, path: str) -> Any:
        """Traverse JSON using dot notation."""
        if path == '.':
            return data
        
        parts = path.strip('.').split('.')
        result = data
        
        for part in parts:
            if part == '':
                continue
            if part.endswith('[]'):
                # Array access
                key = part[:-2]
                result = result.get(key, [])
                if isinstance(result, list):
                    result = list(result)
            elif '[' in part and part.endswith(']'):
                # Indexed access
                key = part.split('[')[0]
                idx = int(part.split('[')[1].rstrip(']'))
                result = result.get(key, [])[idx]
            else:
                result = result.get(part) if isinstance(result, dict) else None
        
        return result
    
    def _apply_operation(self, data: Any, op: Dict) -> Any:
        """Apply a transform operation."""
        op_type = op.get('type')
        
        if op_type == 'map':
            if isinstance(data, list):
                return [self._apply_operation(item, op.get('operation', {})) for item in data]
            return data
        
        elif op_type == 'filter':
            if isinstance(data, list):
                field = op.get('field')
                value = op.get('value')
                return [item for item in data if item.get(field) == value]
            return data
        
        elif op_type == 'select':
            if isinstance(data, dict):
                keys = op.get('keys', [])
                return {k: v for k, v in data.items() if k in keys}
            return data
        
        return data


class XMLParseHandler(TaskHandler):
    """Parse XML files to dictionaries."""
    
    name = "xml_parse"
    category = "data"
    description = "Parse XML files to Python dictionaries"
    requires_filesystem = False
    
    input_schema = {
        'content': {'type': 'string', 'description': 'XML content'},
        'path': {'type': 'string', 'description': 'File path'},
        'attributes_key': {'type': 'string', 'default': '@attributes', 'description': 'Key for attributes'},
        'text_key': {'type': 'string', 'default': '#text', 'description': 'Key for text content'},
    }
    
    output_schema = {
        'data': {'type': 'object', 'description': 'Parsed XML as dictionary'},
        'root_tag': {'type': 'string', 'description': 'Root element tag'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['xml_parse', 'parse_xml', 'xml']:
            return HandlerConfidence.PERFECT.value
        if params.get('path', '').endswith('.xml'):
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        
        try:
            # Get content
            if 'content' in params:
                content = params['content']
            elif 'path' in params:
                with open(params['path'], 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                return HandlerResult(success=False, error="No content or path provided")
            
            attributes_key = params.get('attributes_key', '@attributes')
            text_key = params.get('text_key', '#text')
            
            # Parse XML
            root = ET.fromstring(content)
            
            def element_to_dict(element):
                """Convert XML element to dictionary."""
                result = {}
                
                # Add attributes
                if element.attrib:
                    result[attributes_key] = dict(element.attrib)
                
                # Process children
                children = list(element)
                if children:
                    child_dict = {}
                    for child in children:
                        child_data = element_to_dict(child)
                        if child.tag in child_dict:
                            # Multiple children with same tag
                            if not isinstance(child_dict[child.tag], list):
                                child_dict[child.tag] = [child_dict[child.tag]]
                            child_dict[child.tag].append(child_data)
                        else:
                            child_dict[child.tag] = child_data
                    result.update(child_dict)
                elif element.text and element.text.strip():
                    # Text content
                    if result:  # Has attributes
                        result[text_key] = element.text.strip()
                    else:
                        return element.text.strip()
                
                return result if result else None
            
            data = {root.tag: element_to_dict(root)}
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'data': data,
                    'root_tag': root.tag,
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class ExcelReadHandler(TaskHandler):
    """Read Excel files (.xlsx, .xls)."""
    
    name = "excel_read"
    category = "data"
    description = "Read Excel files to list of dictionaries"
    requires_filesystem = True
    
    input_schema = {
        'path': {'type': 'string', 'required': True, 'description': 'Excel file path'},
        'sheet': {'type': 'string', 'default': None, 'description': 'Sheet name (first sheet if not specified)'},
        'has_header': {'type': 'boolean', 'default': True, 'description': 'First row is header'},
        'start_row': {'type': 'integer', 'default': 1, 'description': 'Start row (1-indexed)'},
        'end_row': {'type': 'integer', 'description': 'End row'},
    }
    
    output_schema = {
        'rows': {'type': 'array', 'description': 'Data rows'},
        'headers': {'type': 'array', 'description': 'Column headers'},
        'sheet_name': {'type': 'string', 'description': 'Sheet name'},
        'sheet_names': {'type': 'array', 'description': 'All sheet names'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['excel_read', 'read_excel', 'xlsx', 'xls']:
            return HandlerConfidence.PERFECT.value
        if params.get('path', '').endswith(('.xlsx', '.xls')):
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        path = params['path']
        sheet = params.get('sheet')
        has_header = params.get('has_header', True)
        start_row = params.get('start_row', 1)
        end_row = params.get('end_row')
        
        try:
            # Try openpyxl for .xlsx
            try:
                from openpyxl import load_workbook
                
                wb = load_workbook(path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                
                if sheet:
                    ws = wb[sheet]
                else:
                    ws = wb.active
                
                rows = []
                headers = []
                
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if i < start_row:
                        continue
                    if end_row and i > end_row:
                        break
                    
                    if i == start_row and has_header:
                        headers = [str(c) if c else f'col_{j}' for j, c in enumerate(row)]
                        continue
                    
                    row_dict = {}
                    for j, val in enumerate(row):
                        key = headers[j] if j < len(headers) else f'col_{j}'
                        row_dict[key] = val
                    rows.append(row_dict)
                
                wb.close()
                
            except ImportError:
                # Fallback: Try xlrd for .xls
                try:
                    import xlrd
                    
                    wb = xlrd.open_workbook(path)
                    sheet_names = wb.sheet_names()
                    
                    if sheet:
                        ws = wb.sheet_by_name(sheet)
                    else:
                        ws = wb.sheet_by_index(0)
                    
                    rows = []
                    headers = []
                    
                    for i in range(ws.nrows):
                        row_num = i + 1
                        if row_num < start_row:
                            continue
                        if end_row and row_num > end_row:
                            break
                        
                        row_values = [ws.cell_value(i, j) for j in range(ws.ncols)]
                        
                        if row_num == start_row and has_header:
                            headers = [str(c) if c else f'col_{j}' for j, c in enumerate(row_values)]
                            continue
                        
                        row_dict = {}
                        for j, val in enumerate(row_values):
                            key = headers[j] if j < len(headers) else f'col_{j}'
                            row_dict[key] = val
                        rows.append(row_dict)
                    
                except ImportError:
                    return HandlerResult(
                        success=False,
                        error="No Excel library available. Install openpyxl or xlrd.",
                    )
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'rows': rows,
                    'headers': headers,
                    'sheet_name': sheet or sheet_names[0],
                    'sheet_names': sheet_names,
                    'row_count': len(rows),
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class PDFExtractHandler(TaskHandler):
    """Extract text from PDF files."""
    
    name = "pdf_extract"
    category = "data"
    description = "Extract text content from PDF files"
    requires_filesystem = True
    
    input_schema = {
        'path': {'type': 'string', 'required': True, 'description': 'PDF file path'},
        'pages': {'type': 'string', 'description': 'Page range (e.g., "1-5, 10, 15-20")'},
        'extract_images': {'type': 'boolean', 'default': False, 'description': 'Extract images'},
        'extract_tables': {'type': 'boolean', 'default': False, 'description': 'Extract tables'},
    }
    
    output_schema = {
        'text': {'type': 'string', 'description': 'Extracted text'},
        'pages': {'type': 'array', 'description': 'Text per page'},
        'page_count': {'type': 'integer', 'description': 'Total pages'},
        'metadata': {'type': 'object', 'description': 'PDF metadata'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['pdf_extract', 'extract_pdf', 'pdf', 'pdf_read']:
            return HandlerConfidence.PERFECT.value
        if params.get('path', '').endswith('.pdf'):
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        path = params['path']
        pages = params.get('pages')
        extract_images = params.get('extract_images', False)
        
        try:
            # Try PyPDF2 / pypdf
            try:
                from pypdf import PdfReader
                
                reader = PdfReader(path)
                page_count = len(reader.pages)
                
                # Parse page range
                page_nums = self._parse_page_range(pages, page_count) if pages else range(page_count)
                
                text_pages = []
                for i in page_nums:
                    page = reader.pages[i]
                    text_pages.append({
                        'page': i + 1,
                        'text': page.extract_text() or '',
                    })
                
                full_text = '\n\n'.join(p['text'] for p in text_pages)
                
                # Metadata
                metadata = {}
                if reader.metadata:
                    metadata = {
                        'title': reader.metadata.get('/Title', ''),
                        'author': reader.metadata.get('/Author', ''),
                        'subject': reader.metadata.get('/Subject', ''),
                        'creator': reader.metadata.get('/Creator', ''),
                        'producer': reader.metadata.get('/Producer', ''),
                    }
                
            except ImportError:
                # Fallback: Try pdfplumber
                try:
                    import pdfplumber
                    
                    with pdfplumber.open(path) as pdf:
                        page_count = len(pdf.pages)
                        page_nums = self._parse_page_range(pages, page_count) if pages else range(page_count)
                        
                        text_pages = []
                        for i in page_nums:
                            page = pdf.pages[i]
                            text_pages.append({
                                'page': i + 1,
                                'text': page.extract_text() or '',
                            })
                        
                        full_text = '\n\n'.join(p['text'] for p in text_pages)
                        metadata = pdf.metadata or {}
                
                except ImportError:
                    return HandlerResult(
                        success=False,
                        error="No PDF library available. Install pypdf or pdfplumber.",
                    )
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'text': full_text,
                    'pages': text_pages,
                    'page_count': page_count,
                    'metadata': metadata,
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )
    
    def _parse_page_range(self, spec: str, max_pages: int) -> List[int]:
        """Parse page range specification like '1-5, 10, 15-20'."""
        result = []
        for part in spec.split(','):
            part = part.strip()
            if '-' in part:
                start, end = map(int, part.split('-'))
                result.extend(range(start - 1, min(end, max_pages)))
            else:
                page = int(part) - 1
                if 0 <= page < max_pages:
                    result.append(page)
        return sorted(set(result))


class DataMergeHandler(TaskHandler):
    """Merge multiple datasets."""
    
    name = "data_merge"
    category = "data"
    description = "Merge multiple datasets by key"
    requires_filesystem = False
    
    input_schema = {
        'sources': {'type': 'array', 'required': True, 'description': 'List of datasets'},
        'key': {'type': 'string', 'required': True, 'description': 'Merge key field'},
        'how': {'type': 'string', 'default': 'inner', 'enum': ['inner', 'left', 'right', 'outer'], 'description': 'Merge type'},
    }
    
    output_schema = {
        'rows': {'type': 'array', 'description': 'Merged data'},
        'row_count': {'type': 'integer', 'description': 'Number of rows'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['data_merge', 'merge', 'join']:
            return HandlerConfidence.PERFECT.value
        if 'sources' in params and 'key' in params:
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        sources = params['sources']
        key = params['key']
        how = params.get('how', 'inner')
        
        try:
            if len(sources) < 2:
                return HandlerResult(success=False, error="Need at least 2 sources to merge")
            
            # Start with first source
            result = {row[key]: dict(row) for row in sources[0] if key in row}
            
            # Merge with subsequent sources
            for source in sources[1:]:
                source_dict = {row[key]: row for row in source if key in row}
                
                if how == 'inner':
                    # Keep only keys in both
                    result = {
                        k: {**result[k], **source_dict[k]}
                        for k in result if k in source_dict
                    }
                elif how == 'left':
                    # Keep all from left
                    for k in result:
                        if k in source_dict:
                            result[k].update(source_dict[k])
                elif how == 'right':
                    # Keep all from right
                    new_result = {}
                    for k in source_dict:
                        if k in result:
                            new_result[k] = {**result[k], **source_dict[k]}
                        else:
                            new_result[k] = source_dict[k]
                    result = new_result
                elif how == 'outer':
                    # Keep all
                    for k in source_dict:
                        if k in result:
                            result[k].update(source_dict[k])
                        else:
                            result[k] = source_dict[k]
            
            rows = list(result.values())
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'rows': rows,
                    'row_count': len(rows),
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class DataFilterHandler(TaskHandler):
    """Filter data with complex conditions."""
    
    name = "data_filter"
    category = "data"
    description = "Filter data with complex conditions"
    requires_filesystem = False
    
    input_schema = {
        'data': {'type': 'any', 'required': True, 'description': 'Input data'},
        'conditions': {'type': 'array', 'required': True, 'description': 'Filter conditions'},
        'mode': {'type': 'string', 'default': 'and', 'enum': ['and', 'or'], 'description': 'Condition combination'},
    }
    
    output_schema = {
        'data': {'type': 'any', 'description': 'Filtered data'},
        'count': {'type': 'integer', 'description': 'Items matching'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['data_filter', 'filter', 'where']:
            return HandlerConfidence.PERFECT.value
        if 'data' in params and 'conditions' in params:
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        data = params['data']
        conditions = params['conditions']
        mode = params.get('mode', 'and')
        
        try:
            if not isinstance(data, list):
                return HandlerResult(
                    success=False,
                    error="Data must be a list",
                )
            
            def matches(item: Dict, condition: Dict) -> bool:
                field = condition.get('field')
                operator = condition.get('operator', '==')
                value = condition.get('value')
                
                item_value = item.get(field)
                
                if operator == '==':
                    return item_value == value
                elif operator == '!=':
                    return item_value != value
                elif operator == '>':
                    return item_value > value
                elif operator == '<':
                    return item_value < value
                elif operator == '>=':
                    return item_value >= value
                elif operator == '<=':
                    return item_value <= value
                elif operator == 'contains':
                    return value in str(item_value)
                elif operator == 'startswith':
                    return str(item_value).startswith(value)
                elif operator == 'endswith':
                    return str(item_value).endswith(value)
                elif operator == 'in':
                    return item_value in value
                elif operator == 'regex':
                    import re
                    return bool(re.match(value, str(item_value)))
                elif operator == 'isnull':
                    return item_value is None or item_value == ''
                elif operator == 'notnull':
                    return item_value is not None and item_value != ''
                
                return False
            
            def matches_all(item: Dict) -> bool:
                results = [matches(item, c) for c in conditions]
                return all(results) if mode == 'and' else any(results)
            
            filtered = [item for item in data if matches_all(item)]
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'data': filtered,
                    'count': len(filtered),
                    'total': len(data),
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class DataValidateHandler(TaskHandler):
    """Validate data against schema."""
    
    name = "data_validate"
    category = "data"
    description = "Validate data against schema"
    requires_filesystem = False
    
    input_schema = {
        'data': {'type': 'any', 'required': True, 'description': 'Data to validate'},
        'schema': {'type': 'object', 'required': True, 'description': 'Validation schema'},
    }
    
    output_schema = {
        'valid': {'type': 'boolean', 'description': 'Is data valid'},
        'errors': {'type': 'array', 'description': 'Validation errors'},
        'warnings': {'type': 'array', 'description': 'Validation warnings'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['data_validate', 'validate', 'validate_schema']:
            return HandlerConfidence.PERFECT.value
        if 'data' in params and 'schema' in params:
            return HandlerConfidence.HIGH.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        data = params['data']
        schema = params['schema']
        
        try:
            errors = []
            warnings = []
            
            # Validate each field
            for field, rules in schema.items():
                value = data.get(field) if isinstance(data, dict) else None
                
                # Required check
                if rules.get('required', False) and value is None:
                    errors.append(f"Missing required field: {field}")
                    continue
                
                if value is None:
                    continue
                
                # Type check
                expected_type = rules.get('type')
                if expected_type:
                    type_map = {
                        'string': str,
                        'integer': int,
                        'number': (int, float),
                        'boolean': bool,
                        'array': list,
                        'object': dict,
                    }
                    expected = type_map.get(expected_type)
                    if expected and not isinstance(value, expected):
                        errors.append(f"Field {field} has wrong type: expected {expected_type}, got {type(value).__name__}")
                
                # Min/max for numbers
                if isinstance(value, (int, float)):
                    if 'min' in rules and value < rules['min']:
                        errors.append(f"Field {field} below minimum: {value} < {rules['min']}")
                    if 'max' in rules and value > rules['max']:
                        errors.append(f"Field {field} above maximum: {value} > {rules['max']}")
                
                # Min/max length for strings
                if isinstance(value, str):
                    if 'min_length' in rules and len(value) < rules['min_length']:
                        errors.append(f"Field {field} too short: {len(value)} < {rules['min_length']}")
                    if 'max_length' in rules and len(value) > rules['max_length']:
                        errors.append(f"Field {field} too long: {len(value)} > {rules['max_length']}")
                
                # Pattern for strings
                if isinstance(value, str) and 'pattern' in rules:
                    import re
                    if not re.match(rules['pattern'], value):
                        errors.append(f"Field {field} doesn't match pattern: {rules['pattern']}")
                
                # Enum
                if 'enum' in rules:
                    if value not in rules['enum']:
                        errors.append(f"Field {field} not in allowed values: {rules['enum']}")
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'valid': len(errors) == 0,
                    'errors': errors,
                    'warnings': warnings,
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )


class DataAggregateHandler(TaskHandler):
    """Aggregate data with group-by operations."""
    
    name = "data_aggregate"
    category = "data"
    description = "Aggregate data with group-by operations"
    requires_filesystem = False
    
    input_schema = {
        'data': {'type': 'array', 'required': True, 'description': 'Input data'},
        'group_by': {'type': 'string', 'description': 'Field to group by'},
        'aggregations': {'type': 'array', 'description': 'Aggregation operations'},
    }
    
    output_schema = {
        'data': {'type': 'array', 'description': 'Aggregated data'},
        'groups': {'type': 'integer', 'description': 'Number of groups'},
    }
    
    def can_handle(self, params: Dict[str, Any]) -> float:
        if params.get('task') in ['data_aggregate', 'aggregate', 'group_by', 'groupby']:
            return HandlerConfidence.PERFECT.value
        return 0.0
    
    def execute(self, params: Dict[str, Any]) -> HandlerResult:
        start_time = time.time()
        data = params['data']
        group_by = params.get('group_by')
        aggregations = params.get('aggregations', [])
        
        try:
            if group_by:
                # Group data
                groups = {}
                for item in data:
                    key = item.get(group_by)
                    if key not in groups:
                        groups[key] = []
                    groups[key].append(item)
                
                # Aggregate each group
                result = []
                for key, items in groups.items():
                    row = {group_by: key, '_count': len(items)}
                    
                    for agg in aggregations:
                        field = agg.get('field')
                        op = agg.get('operation') or agg.get('op')
                        as_field = agg.get('as') or f'{op}_{field}'
                        
                        values = [i.get(field) for i in items if i.get(field) is not None]
                        
                        if op == 'sum':
                            row[as_field] = sum(values) if values else 0
                        elif op == 'avg' or op == 'mean':
                            row[as_field] = sum(values) / len(values) if values else 0
                        elif op == 'min':
                            row[as_field] = min(values) if values else None
                        elif op == 'max':
                            row[as_field] = max(values) if values else None
                        elif op == 'count':
                            row[as_field] = len(values)
                        elif op == 'first':
                            row[as_field] = values[0] if values else None
                        elif op == 'last':
                            row[as_field] = values[-1] if values else None
                    
                    result.append(row)
                
            else:
                # Aggregate all data
                row = {'_count': len(data)}
                
                for agg in aggregations:
                    field = agg.get('field')
                    op = agg.get('operation') or agg.get('op')
                    as_field = agg.get('as') or f'{op}_{field}'
                    
                    values = [i.get(field) for i in data if i.get(field) is not None]
                    
                    if op == 'sum':
                        row[as_field] = sum(values) if values else 0
                    elif op == 'avg' or op == 'mean':
                        row[as_field] = sum(values) / len(values) if values else 0
                    elif op == 'min':
                        row[as_field] = min(values) if values else None
                    elif op == 'max':
                        row[as_field] = max(values) if values else None
                    elif op == 'count':
                        row[as_field] = len(values)
                
                result = [row]
            
            duration_ms = (time.time() - start_time) * 1000
            
            return HandlerResult(
                success=True,
                data={
                    'data': result,
                    'groups': len(result),
                },
                duration_ms=duration_ms,
            )
            
        except Exception as e:
            return HandlerResult(
                success=False,
                error=str(e),
                duration_ms=(time.time() - start_time) * 1000,
            )
